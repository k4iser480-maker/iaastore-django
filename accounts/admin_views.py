import json
import django
import sys
from datetime import datetime, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum, Count, F, Q, ExpressionWrapper, FloatField
from orders.services.referral_service import ReferralService
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.conf import settings

from accounts.models import Account, Role, Permission, ActivityLog
from orders.models import Order, OrderProduct, Payment, Transportista
from store.models import Product, ReviewRating, Wishlist
from category.models import Category
from accounts.forms import AdminProfileForm
from accounts.decorators import has_permission, admin_panel_access_required


# ==========================================
# HELPER: Log activity
# ==========================================
def log_activity(request, action, module, description):
    ip = request.META.get('HTTP_X_FORWARDED_FOR')
    if ip:
        ip = ip.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')
    ActivityLog.objects.create(
        user=request.user,
        action=action,
        module=module,
        description=description,
        ip_address=ip,
    )


# ==========================================
# DASHBOARD
# ==========================================
@admin_panel_access_required
def admin_dashboard(request):
    if request.user.is_transportista and not request.user.is_superadmin:
        return redirect('admin_orders')
        
    now = timezone.now()
    thirty_days_ago = now - timedelta(days=30)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # --- Stat cards ---
    all_orders = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"])
    total_revenue = all_orders.aggregate(s=Sum('order_total'))['s'] or 0
    total_orders = all_orders.count()
    avg_order_value = (total_revenue / total_orders) if total_orders > 0 else 0
    completed_orders = all_orders.filter(status='Completed').count()
    total_customers = Account.objects.filter(is_admin=False, is_staff=False).count()
    active_customers = Account.objects.filter(is_active=True, is_admin=False, is_staff=False).count()

    # --- Status counts ---
    status_completed = all_orders.filter(status='Completed').count()
    status_accepted = all_orders.filter(status='Accepted').count()
    status_new = all_orders.filter(status='New').count()
    status_cancelled = all_orders.filter(status='Cancelled').count()

    # --- Daily sales (last 30 days) ---
    daily_sales = (
        all_orders
        .filter(created_at__gte=thirty_days_ago)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(total=Sum('order_total'))
        .order_by('day')
    )
    daily_labels = []
    daily_data = []
    for entry in daily_sales:
        daily_labels.append(entry['day'].strftime('%d %b'))
        daily_data.append(float(entry['total']))

    # If no data, add placeholder
    if not daily_labels:
        for i in range(30, 0, -1):
            d = (now - timedelta(days=i)).strftime('%d %b')
            daily_labels.append(d)
            daily_data.append(0)

    # --- Recent orders ---
    recent_orders = all_orders.order_by('-created_at')[:8]

    # --- Top products ---
    top_products = (
        OrderProduct.objects
        .filter(ordered=True, order__is_test=False)
        .values('product__product_name', 'product__images')
        .annotate(total_sold=Sum('quantity'), revenue=Sum(F('quantity') * F('product_price')))
        .order_by('-total_sold')[:5]
    )

    # --- Monthly progress ---
    monthly_orders_qs = all_orders.filter(created_at__gte=month_start)
    monthly_revenue = monthly_orders_qs.aggregate(s=Sum('order_total'))['s'] or 0
    monthly_orders = monthly_orders_qs.count()
    monthly_customers = Account.objects.filter(date_joined__gte=month_start, is_admin=False, is_staff=False).count()

    monthly_goal = 5000
    orders_goal = 50
    customers_goal = 20

    revenue_pct = min(100, (monthly_revenue / monthly_goal * 100) if monthly_goal else 0)
    orders_pct = min(100, (monthly_orders / orders_goal * 100) if orders_goal else 0)
    customers_pct = min(100, (monthly_customers / customers_goal * 100) if customers_goal else 0)

    # --- Pending orders count for sidebar badge ---
    pending_orders_count = all_orders.filter(status__in=['New', 'Processing']).count()

    context = {
        'active_page': 'dashboard',
        'total_revenue': total_revenue,
        'avg_order_value': avg_order_value,
        'total_orders': total_orders,
        'completed_orders': completed_orders,
        'total_customers': total_customers,
        'active_customers': active_customers,
        'status_completed': status_completed,
        'status_accepted': status_accepted,
        'status_new': status_new,
        'status_cancelled': status_cancelled,
        'daily_labels': json.dumps(daily_labels),
        'daily_data': json.dumps(daily_data),
        'recent_orders': recent_orders,
        'top_products': top_products,
        'monthly_revenue': monthly_revenue,
        'monthly_orders': monthly_orders,
        'monthly_customers': monthly_customers,
        'monthly_goal': monthly_goal,
        'orders_goal': orders_goal,
        'customers_goal': customers_goal,
        'revenue_pct': revenue_pct,
        'orders_pct': orders_pct,
        'customers_pct': customers_pct,
        'pending_orders_count': pending_orders_count,
    }
    return render(request, 'admin_panel/dashboard.html', context)

# ==========================================
# ORDER STATE MACHINE
# ==========================================
VALID_TRANSITIONS = {
    'New':             ['Processing', 'Cancelled'],
    'Processing':      ['Assigned', 'Completed', 'Cancelled'],
    'Assigned':        ['Picked Up', 'In Transit', 'Cancelled'],
    'Picked Up':       ['In Transit', 'Cancelled'],
    'In Transit':      ['Nearby', 'Delivered', 'Failed Attempt', 'Cancelled'],
    'Nearby':          ['Delivered', 'Failed Attempt', 'Cancelled'],
    'Delivered':       ['Completed'],
    'Completed':       [],
    'Cancelled':       [],
    'Failed Attempt':  ['Assigned', 'Cancelled'],
}

# Status display labels
STATUS_LABELS = {
    'New': 'Nuevo',
    'Processing': 'Procesando',
    'Assigned': 'Asignado',
    'Picked Up': 'Recogido',
    'In Transit': 'En Camino',
    'Nearby': 'Cerca del Destino',
    'Delivered': 'Entregado',
    'Completed': 'Completado',
    'Cancelled': 'Cancelado',
    'Failed Attempt': 'Intento Fallido',
}



# ==========================================
# ORDERS
# ==========================================
@has_permission('view_orders')
def admin_orders(request):
    show_sandbox = request.GET.get('type') == 'sandbox'
    if show_sandbox:
        orders = Order.objects.sandbox().filter(payment_status__in=["review", "paid", "failed"]).select_related('payment', 'transportista').order_by('-created_at')
    else:
        orders = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"]).select_related('payment', 'transportista').order_by('-created_at')

    is_transportista = request.user.is_transportista and not request.user.is_superadmin
    if is_transportista:
        return redirect('transportista:transportista_dashboard')

    filter_status = request.GET.get('status', '')
    if filter_status:
        orders = orders.filter(status=filter_status)

    transportistas = Transportista.objects.filter(disponible=True)
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['Processing', 'Assigned']).count()

    context = {
        'active_page': 'orders',
        'orders': orders,
        'filter_status': filter_status,
        'transportistas': transportistas,
        'pending_orders_count': pending_orders_count,
        'is_transportista': is_transportista,
        'valid_transitions': json.dumps(VALID_TRANSITIONS),
        'status_labels': json.dumps(STATUS_LABELS),
        'show_sandbox': show_sandbox,
    }
    return render(request, 'admin_panel/orders.html', context)


@admin_panel_access_required
def admin_update_order_status(request, order_id):
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        old_status = order.status
        new_status = request.POST.get('status')
        transportista_id = request.POST.get('transportista_id')
        force_reason = request.POST.get('force_reason', '').strip()
        new_payment_status = request.POST.get('payment_status')

        # --- Verificación de Pago Manual ---
        if new_payment_status and order.payment_status == 'review':
            if new_payment_status == 'paid':
                order.payment_status = 'paid'
                order.status = 'Processing'
                if order.payment:
                    order.payment.status = 'COMPLETED'
                    order.payment.save()
                order.save()
                
                # Process referral reward
                ReferralService.reward_referral(order)
                
                messages.success(request, f'Pago aprobado. El pedido #{order.order_number} ha pasado a estado "Procesando".')
                log_activity(request, 'PAYMENT', 'Pedidos', f'Aprobó el pago manual del pedido #{order.order_number}.')
            elif new_payment_status == 'failed':
                order.payment_status = 'failed'
                order.status = 'Cancelled'
                if order.payment:
                    order.payment.status = 'FAILED'
                    order.payment.save()
                order.save()
                messages.error(request, f'Pago rechazado. El pedido #{order.order_number} ha sido cancelado.')
                log_activity(request, 'PAYMENT', 'Pedidos', f'Rechazó el pago manual del pedido #{order.order_number}.')
            return redirect('admin_orders')

        # --- Bloqueo: pedido no pagado ---
        if order.payment_status != "paid" and order.status != 'Cancelled':
            if new_status or transportista_id:
                messages.error(request, 'No se puede cambiar la logística de un pedido sin pago confirmado.')
            return redirect('admin_orders')

        # --- Validacion de transicion ---
        is_transportista_assigning = transportista_id and not new_status
        if new_status and new_status != old_status and not is_transportista_assigning:
            allowed = VALID_TRANSITIONS.get(old_status, [])

            if new_status not in allowed:
                # Superadmin puede forzar con razon
                if request.user.is_superadmin and force_reason:
                    order.status = new_status
                    order.is_exception = True
                    order.exception_reason = f'Transicion forzada: {old_status} -> {new_status}. Razon: {force_reason}'
                    log_activity(request, 'STATUS', 'Pedidos',
                                 f'FORZÓ transicion del pedido #{order.order_number}: "{old_status}" -> "{new_status}". Razon: {force_reason}.')
                elif request.user.is_superadmin and not force_reason:
                    messages.error(request, f'Transicion "{old_status}" -> "{new_status}" no permitida. Como superadmin puedes forzarla proporcionando una razon.')
                    return redirect('admin_orders')
                else:
                    messages.error(request, f'Transicion no permitida: "{old_status}" -> "{new_status}".')
                    return redirect('admin_orders')
            else:
                order.status = new_status

        # --- Asignacion de transportista ---
        if transportista_id:
            try:
                transportista = Transportista.objects.get(id=transportista_id)
                old_transportista = order.transportista

                order.transportista = transportista

                # Logic to auto set status to asignado if assigned for the first time
                if old_transportista is None and transportista is not None:
                    # Validar que la transicion pendiente -> asignado es valida
                    if order.status in ['Processing'] or (new_status and order.status == new_status):
                        order.status = 'Assigned'

                    # Send email
                    from django.template.loader import render_to_string
                    from django.core.mail import EmailMessage
                    from django.contrib.sites.shortcuts import get_current_site
                    current_site = get_current_site(request)

                    mail_subject = 'Nuevo pedido asignado'
                    message = render_to_string('orders/transportista_email.html', {
                        'transportista': transportista,
                        'order': order,
                        'domain': current_site.domain,
                    })
                    try:
                        send_email = EmailMessage(mail_subject, message, to=[transportista.email_notificaciones or transportista.user.email])
                        send_email.content_subtype = "html"
                        send_email.send()
                    except Exception as e:
                        pass
            except Transportista.DoesNotExist:
                pass

        # --- Logica de disponibilidad del transportista ---
        if order.transportista:
            t = order.transportista

            if order.status in ['Picked Up', 'In Transit', 'Nearby']:
                t.disponible = False
                t.save()

            elif order.status in ['Delivered', 'Completed', 'Cancelled']:
                if order.status == 'Delivered' and not order.fecha_entrega:
                    order.fecha_entrega = timezone.now()

                # Check si tiene otros pedidos activos
                has_active_orders = Order.objects.filter(
                    transportista=t,
                    status__in=['Assigned', 'Picked Up', 'In Transit', 'Nearby']
                ).exclude(id=order.id).exists()

                if not has_active_orders:
                    t.disponible = True
                    t.save()

        order.save()

        # --- Auto-create DeliveryCheckpoint for delivery-related status changes ---
        from orders.models import DeliveryCheckpoint, DeliveryStatus
        ADMIN_STATUS_TO_DELIVERY = {
            'Picked Up': DeliveryStatus.PICKED_UP,
            'In Transit': DeliveryStatus.IN_TRANSIT,
            'Nearby': DeliveryStatus.NEARBY,
            'Delivered': DeliveryStatus.DELIVERED,
            'Failed Attempt': DeliveryStatus.FAILED_ATTEMPT,
        }
        delivery_status = ADMIN_STATUS_TO_DELIVERY.get(order.status)
        if delivery_status and order.status != old_status:
            DeliveryCheckpoint.objects.create(
                order=order,
                status=delivery_status,
                note=f'Estado cambiado desde panel admin',
                created_by=request.user,
            )

        if not order.is_exception:
            log_activity(request, 'STATUS', 'Pedidos',
                         f'Actualizó el pedido #{order.order_number}. Estado: "{old_status}" -> "{order.status}".')
    return redirect('admin_orders')



@has_permission('delete_orders')
def admin_delete_order(request, order_id):
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        order_number = order.order_number
        order.delete()
        log_activity(request, 'DELETE', 'Pedidos', f'Eliminó el pedido #{order_number}.')
        messages.success(request, f'Pedido #{order_number} eliminado correctamente.')
    return redirect('admin_orders')


@has_permission('delete_orders')
def admin_clear_orders(request):
    if request.method == 'POST':
        password = request.POST.get('password')
        if request.user.check_password(password):
            count = Order.objects.count()
            OrderProduct.objects.all().delete()
            Payment.objects.all().delete()
            Order.objects.all().delete()
            log_activity(request, 'DELETE', 'Pedidos', f'Eliminó TODOS los pedidos ({count} registros).')
            messages.success(request, f'Se eliminaron {count} pedidos correctamente.')
        else:
            messages.error(request, 'Contraseña incorrecta. No se pudieron eliminar los registros.')
    return redirect('admin_orders')


# ==========================================
# PRODUCTS
# ==========================================
@has_permission('view_inventory')
def admin_products(request):
    products = Product.objects.select_related('Category').order_by('-created_date')
    categories = Category.objects.all()
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()

    context = {
        'active_page': 'products',
        'products': products,
        'categories': categories,
        'pending_orders_count': pending_orders_count,
    }
    return render(request, 'admin_panel/products.html', context)


@admin_panel_access_required
@has_permission('edit_inventory')
def force_update_bcv(request):
    try:
        from django.core.management import call_command
        call_command('update_bcv_rate')
        messages.success(request, 'La tasa del BCV se ha actualizado correctamente.')
        log_activity(request, 'Actualización', 'ExchangeRate', 'Forzó la actualización manual de la tasa BCV.')
    except Exception as e:
        messages.error(request, f'Error al actualizar la tasa del BCV: {str(e)}')
    return redirect('admin_dashboard')


@admin_panel_access_required
def admin_toggle_product(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=product_id)
        product.is_available = not product.is_available
        product.save()
        status_text = "activó" if product.is_available else "desactivó"
        log_activity(request, 'TOGGLE', 'Productos',
                     f'{status_text} el producto "{product.product_name}" (ID: {product.id}).')
    return redirect('admin_products')


@has_permission('create_inventory')
def admin_create_product(request):
    from store.forms import ProductForm
    from django.utils.text import slugify
    from store.models import ProductFeature

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.slug = slugify(product.product_name)
            # Ensure unique slug
            base_slug = product.slug
            counter = 1
            while Product.objects.filter(slug=product.slug).exists():
                product.slug = f"{base_slug}-{counter}"
                counter += 1
            product.save()
            
            # Save features
            feature_names = request.POST.getlist('feature_name[]')
            feature_values = request.POST.getlist('feature_value[]')
            for name, value in zip(feature_names, feature_values):
                if name.strip() and value.strip():
                    ProductFeature.objects.create(product=product, name=name.strip(), value=value.strip())
                    
            log_activity(request, 'CREATE', 'Productos',
                         f'Creó el producto "{product.product_name}" (Categoría: {product.Category.category_name}, Precio: ${product.price}, Stock: {product.stock}).')
            messages.success(request, 'Producto creado exitosamente.')
            return redirect('admin_products')
    else:
        form = ProductForm()

    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()
    return render(request, 'admin_panel/product_form.html', {
        'active_page': 'products',
        'form': form,
        'action': 'Crear',
        'pending_orders_count': pending_orders_count,
    })


@has_permission('edit_inventory')
def admin_edit_product(request, product_id):
    from store.forms import ProductForm
    from django.utils.text import slugify
    from store.models import ProductFeature

    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            updated = form.save(commit=False)
            updated.slug = slugify(updated.product_name)
            base_slug = updated.slug
            counter = 1
            while Product.objects.filter(slug=updated.slug).exclude(id=product.id).exists():
                updated.slug = f"{base_slug}-{counter}"
                counter += 1
            updated.save()
            
            # Update features
            updated.features.all().delete()
            feature_names = request.POST.getlist('feature_name[]')
            feature_values = request.POST.getlist('feature_value[]')
            for name, value in zip(feature_names, feature_values):
                if name.strip() and value.strip():
                    ProductFeature.objects.create(product=updated, name=name.strip(), value=value.strip())
                    
            log_activity(request, 'UPDATE', 'Productos',
                         f'Editó el producto "{updated.product_name}" (ID: {updated.id}).')
            messages.success(request, 'Producto actualizado exitosamente.')
            return redirect('admin_products')
    else:
        form = ProductForm(instance=product)

    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()
    return render(request, 'admin_panel/product_form.html', {
        'active_page': 'products',
        'form': form,
        'product': product,
        'action': 'Editar',
        'pending_orders_count': pending_orders_count,
    })


@has_permission('delete_inventory')
def admin_delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product_name = product.product_name
        product.delete()
        log_activity(request, 'DELETE', 'Productos',
                     f'Eliminó el producto "{product_name}" (ID: {product_id}).')
        messages.success(request, 'Producto eliminado.')
    return redirect('admin_products')


@has_permission('delete_inventory')
def admin_clear_products(request):
    if request.method == 'POST':
        password = request.POST.get('password')
        if request.user.check_password(password):
            count = Product.objects.count()
            Product.objects.all().delete()
            log_activity(request, 'DELETE', 'Productos', f'Eliminó TODOS los productos ({count} registros).')
            messages.success(request, f'Se eliminaron {count} productos correctamente.')
        else:
            messages.error(request, 'Contraseña incorrecta. No se pudieron eliminar los registros.')
    return redirect('admin_products')


# ==========================================
# CUSTOMERS (legacy view, kept for compatibility)
# ==========================================
@has_permission('view_users')
def admin_customers(request):
    customers = (
        Account.objects
        .filter(is_admin=False, is_staff=False)
        .annotate(
            order_count=Count('order', filter=Q(order__payment_status__in=["review", "paid", "failed"], order__is_test=False)),
            total_spent=Sum('order__order_total', filter=Q(order__payment_status__in=["review", "paid", "failed"], order__is_test=False))
        )
        .order_by('-date_joined')
    )
    # Default total_spent to 0
    for c in customers:
        if c.total_spent is None:
            c.total_spent = 0

    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()

    context = {
        'active_page': 'customers',
        'customers': customers,
        'pending_orders_count': pending_orders_count,
    }
    return render(request, 'admin_panel/customers.html', context)


# ==========================================
# SETTINGS
# ==========================================
@admin_panel_access_required
def admin_settings(request):
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()

    context = {
        'active_page': 'settings',
        'django_version': django.get_version(),
        'python_version': f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
        'db_products': Product.objects.count(),
        'db_orders': Order.objects.real().filter(payment_status__in=["review", "paid", "failed"]).count(),
        'db_users': Account.objects.filter(is_admin=False, is_staff=False).count(),
        'db_reviews': ReviewRating.objects.count(),
        'db_categories': Category.objects.count(),
        'db_wishlist': Wishlist.objects.count(),
        'pending_orders_count': pending_orders_count,
    }
    return render(request, 'admin_panel/settings.html', context)


# ==========================================
# PROFILE
# ==========================================
@admin_panel_access_required
def admin_profile(request):
    if request.method == 'POST':
        form = AdminProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            log_activity(request, 'UPDATE', 'Perfil',
                         f'Actualizó su perfil de administrador.')
            messages.success(request, 'Tu perfil ha sido actualizado correctamente.')
            return redirect('admin_profile')
    else:
        form = AdminProfileForm(instance=request.user)
    
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()

    context = {
        'active_page': 'profile',
        'form': form,
        'pending_orders_count': pending_orders_count,
    }
    return render(request, 'admin_panel/profile.html', context)


# ==========================================
# ROLES MANAGEMENT
# ==========================================
@has_permission('view_roles')
def admin_roles(request):
    roles = Role.objects.annotate(user_count=Count('account')).all()
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()
    context = {
        'active_page': 'roles',
        'roles': roles,
        'pending_orders_count': pending_orders_count,
    }
    return render(request, 'admin_panel/roles.html', context)

@has_permission('create_roles')
def admin_create_role(request):
    permissions = Permission.objects.all().order_by('module')
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        selected_perms = request.POST.getlist('permissions')
        
        if Role.objects.filter(name=name).exists():
            messages.error(request, 'Ya existe un rol con ese nombre.')
        else:
            role = Role.objects.create(name=name, description=description)
            role.permissions.set(selected_perms)
            perm_names = Permission.objects.filter(id__in=selected_perms).values_list('codename', flat=True)
            log_activity(request, 'CREATE', 'Roles',
                         f'Creó el rol "{name}" con permisos: {", ".join(perm_names)}.')
            messages.success(request, 'Rol creado exitosamente.')
            return redirect('admin_roles')
            
    # Group permissions by module for the matrix
    modules = {}
    for p in permissions:
        if p.module not in modules:
            modules[p.module] = []
        modules[p.module].append(p)
        
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()
    return render(request, 'admin_panel/role_form.html', {
        'active_page': 'roles',
        'modules': modules,
        'action': 'Crear',
        'pending_orders_count': pending_orders_count
    })

@has_permission('edit_roles')
def admin_edit_role(request, role_id):
    role = get_object_or_404(Role, id=role_id)
    permissions = Permission.objects.all().order_by('module')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        selected_perms = request.POST.getlist('permissions')
        
        if Role.objects.filter(name=name).exclude(id=role.id).exists():
            messages.error(request, 'Ya existe un rol con ese nombre.')
        else:
            old_name = role.name
            role.name = name
            role.description = description
            role.permissions.set(selected_perms)
            role.save()
            perm_names = Permission.objects.filter(id__in=selected_perms).values_list('codename', flat=True)
            log_activity(request, 'UPDATE', 'Roles',
                         f'Editó el rol "{old_name}" → "{name}". Permisos actualizados: {", ".join(perm_names)}.')
            messages.success(request, 'Rol actualizado exitosamente.')
            return redirect('admin_roles')
            
    modules = {}
    for p in permissions:
        if p.module not in modules:
            modules[p.module] = []
        modules[p.module].append(p)
        
    role_perms = role.permissions.values_list('id', flat=True)
        
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()
    return render(request, 'admin_panel/role_form.html', {
        'active_page': 'roles',
        'role': role,
        'modules': modules,
        'role_perms': role_perms,
        'action': 'Editar',
        'pending_orders_count': pending_orders_count
    })

@has_permission('delete_roles')
def admin_delete_role(request, role_id):
    role = get_object_or_404(Role, id=role_id)
    if request.method == 'POST':
        role_name = role.name
        role.delete()
        log_activity(request, 'DELETE', 'Roles',
                     f'Eliminó el rol "{role_name}".')
        messages.success(request, 'Rol eliminado.')
    return redirect('admin_roles')

@has_permission('create_roles')
def admin_clone_role(request, role_id):
    role = get_object_or_404(Role, id=role_id)
    if request.method == 'POST':
        new_name = role.name + " (Copia)"
        i = 1
        while Role.objects.filter(name=new_name).exists():
            new_name = f"{role.name} (Copia {i})"
            i += 1
            
        new_role = Role.objects.create(name=new_name, description=role.description)
        new_role.permissions.set(role.permissions.all())
        log_activity(request, 'CREATE', 'Roles',
                     f'Duplicó el rol "{role.name}" como "{new_name}".')
        messages.success(request, 'Rol duplicado exitosamente. Ahora puedes editarlo.')
        return redirect('admin_edit_role', role_id=new_role.id)
    return redirect('admin_roles')


# ==========================================
# USERS MANAGEMENT
# ==========================================
@has_permission('view_users')
def admin_users(request):
    users = Account.objects.all().order_by('-date_joined')
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()
    context = {
        'active_page': 'users',
        'users': users,
        'pending_orders_count': pending_orders_count,
    }
    return render(request, 'admin_panel/users.html', context)

@has_permission('create_users')
def admin_create_user(request):
    from accounts.forms import RegistrationForm
    roles = Role.objects.all()
    
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            username = email.split("@")[0]
            
            user = Account.objects.create_user(first_name=first_name, last_name=last_name, email=email, username=username, password=password)
            user.is_active = True # auto activate for admin created users
            
            # Only superadmin can assign roles
            if request.user.is_superadmin:
                selected_roles = request.POST.getlist('roles')
                if selected_roles:
                    user.is_staff = True
                    user.save()
                    user.roles.set(selected_roles)
                    role_names = Role.objects.filter(id__in=selected_roles).values_list('name', flat=True)
                    log_activity(request, 'CREATE', 'Usuarios',
                                 f'Creó al usuario "{first_name} {last_name}" ({email}) con roles: {", ".join(role_names)}.')
                else:
                    user.save()
                    log_activity(request, 'CREATE', 'Usuarios',
                                 f'Creó al usuario "{first_name} {last_name}" ({email}) sin roles asignados.')
            else:
                user.save()
                log_activity(request, 'CREATE', 'Usuarios',
                             f'Creó al usuario "{first_name} {last_name}" ({email}).')
                
            messages.success(request, 'Usuario creado exitosamente.')
            return redirect('admin_users')
    else:
        form = RegistrationForm()
        
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()
    return render(request, 'admin_panel/user_form.html', {
        'active_page': 'users',
        'form': form,
        'roles': roles,
        'action': 'Crear',
        'pending_orders_count': pending_orders_count
    })

@has_permission('edit_users')
def admin_edit_user(request, user_id):
    edit_user = get_object_or_404(Account, id=user_id)
    roles = Role.objects.all()
    
    # Superadmin protection
    if edit_user.is_superadmin and not request.user.is_superadmin:
        messages.error(request, 'No tienes permisos para editar a un Super Administrador.')
        return redirect('admin_users')
        
    if request.method == 'POST':
        changes = []
        old_first = edit_user.first_name
        old_last = edit_user.last_name
        
        edit_user.first_name = request.POST.get('first_name', edit_user.first_name)
        edit_user.last_name = request.POST.get('last_name', edit_user.last_name)
        edit_user.phone_number = request.POST.get('phone_number', edit_user.phone_number)
        
        if old_first != edit_user.first_name or old_last != edit_user.last_name:
            changes.append(f'nombre: "{old_first} {old_last}" → "{edit_user.first_name} {edit_user.last_name}"')
        
        is_active = request.POST.get('is_active') == 'on'
        # Prevent deactivating superadmin
        if edit_user.is_superadmin:
            is_active = True
        
        if edit_user.is_active != is_active:
            changes.append(f'estado: {"Activo" if is_active else "Inactivo"}')
        edit_user.is_active = is_active
        
        # Only superadmin can change roles
        if request.user.is_superadmin:
            old_roles = set(edit_user.roles.values_list('name', flat=True))
            selected_roles = request.POST.getlist('roles')
            if selected_roles:
                edit_user.is_staff = True
            else:
                if not edit_user.is_superadmin:
                    edit_user.is_staff = False
            edit_user.save()
            edit_user.roles.set(selected_roles)
            new_roles = set(edit_user.roles.values_list('name', flat=True))
            if old_roles != new_roles:
                changes.append(f'roles: [{", ".join(old_roles) or "ninguno"}] → [{", ".join(new_roles) or "ninguno"}]')
        else:
            edit_user.save()

        desc = f'Editó al usuario "{edit_user.email}".'
        if changes:
            desc += ' Cambios: ' + '; '.join(changes) + '.'
        log_activity(request, 'UPDATE', 'Usuarios', desc)
        messages.success(request, 'Usuario actualizado exitosamente.')
        return redirect('admin_users')
        
    user_roles = edit_user.roles.values_list('id', flat=True)
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()
    return render(request, 'admin_panel/user_form.html', {
        'active_page': 'users',
        'edit_user': edit_user,
        'roles': roles,
        'user_roles': user_roles,
        'action': 'Editar',
        'pending_orders_count': pending_orders_count
    })

@has_permission('delete_users')
def admin_delete_user(request, user_id):
    target_user = get_object_or_404(Account, id=user_id)
    if request.method == 'POST':
        if target_user.is_superadmin:
            messages.error(request, 'No puedes eliminar a un Super Administrador.')
        elif target_user == request.user:
            messages.error(request, 'No puedes eliminarte a ti mismo.')
        else:
            user_email = target_user.email
            user_name = f"{target_user.first_name} {target_user.last_name}"
            target_user.delete()
            log_activity(request, 'DELETE', 'Usuarios',
                         f'Eliminó al usuario "{user_name}" ({user_email}).')
            messages.success(request, 'Usuario eliminado.')
    return redirect('admin_users')


# ==========================================
# ACTIVITY LOG (Superadmin only)
# ==========================================
@admin_panel_access_required
def admin_activity_log(request):
    if not request.user.is_superadmin:
        messages.error(request, 'Solo el Super Administrador puede ver el registro de actividad.')
        return redirect('admin_dashboard')

    logs = ActivityLog.objects.select_related('user').all()

    # Filters
    filter_module = request.GET.get('module', '')
    filter_action = request.GET.get('action', '')
    filter_user = request.GET.get('user_id', '')

    if filter_module:
        logs = logs.filter(module=filter_module)
    if filter_action:
        logs = logs.filter(action=filter_action)
    if filter_user:
        logs = logs.filter(user_id=filter_user)

    # Get unique modules and users for filter dropdowns
    all_modules = ActivityLog.objects.values_list('module', flat=True).distinct().order_by('module')
    all_users = Account.objects.filter(is_staff=True).order_by('first_name')

    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()

    context = {
        'active_page': 'logs',
        'logs': logs[:200],  # Limit to last 200 entries
        'filter_module': filter_module,
        'filter_action': filter_action,
        'filter_user': filter_user,
        'all_modules': all_modules,
        'all_users': all_users,
        'action_choices': ActivityLog.ACTION_CHOICES,
        'pending_orders_count': pending_orders_count,
    }
    return render(request, 'admin_panel/activity_log.html', context)


# ==========================================
# TRANSPORTISTAS MANAGEMENT
# ==========================================
@has_permission('view_orders')
def admin_transportistas(request):
    transportistas = Transportista.objects.select_related('user').all()
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing', 'Assigned']).count()
    context = {
        'active_page': 'transportistas',
        'transportistas': transportistas,
        'pending_orders_count': pending_orders_count,
    }
    return render(request, 'admin_panel/transportistas.html', context)

@has_permission('view_orders')
def admin_create_transportista(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        telefono = request.POST.get('telefono')
        vehiculo = request.POST.get('vehiculo', '')
        email_notificaciones = request.POST.get('email_notificaciones', '')
        
        try:
            user = Account.objects.get(id=user_id)
            if hasattr(user, 'transportista'):
                messages.error(request, f'El usuario {user.email} ya es un transportista.')
            else:
                Transportista.objects.create(
                    user=user,
                    telefono=telefono,
                    vehiculo=vehiculo,
                    email_notificaciones=email_notificaciones
                )
                messages.success(request, 'Transportista creado exitosamente.')
                log_activity(request, 'CREATE', 'Transportistas', f'Registró a {user.email} como transportista.')
                return redirect('admin_transportistas')
        except Account.DoesNotExist:
            messages.error(request, 'Usuario inválido.')

    users = Account.objects.filter(is_admin=False, is_staff=False)
    # Filter users that don't have transportista profile
    available_users = [u for u in users if not hasattr(u, 'transportista')]
    
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()
    return render(request, 'admin_panel/transportista_form.html', {
        'active_page': 'transportistas',
        'users': available_users,
        'pending_orders_count': pending_orders_count
    })


# ==========================================
# BACKUPS MANAGEMENT
# ==========================================
import os
from pathlib import Path as FilePath
from django.http import FileResponse, JsonResponse
from django.core import serializers
from django.db import transaction
from carts.models import Cart, CartItem
from ecosite.models import Ticket

BACKUP_DIR = os.path.join(settings.MEDIA_ROOT, 'backups')
SCHEMA_VERSION = '1.0'

# Map of module names -> model classes
BACKUP_MODULES = {
    'Usuarios': [Account, 'accounts.ShippingAddress'],
    'Roles': [Role, Permission],
    'Categorías': [Category],
    'Productos': [Product, 'store.ProductFeature', ReviewRating, Wishlist, 'store.ExchangeRate'],
    'Pedidos': [Order, 'orders.OrderProduct', Payment, Transportista],
    'Carritos': [Cart, CartItem],
    'Tickets': [Ticket],
    'Actividad': [ActivityLog],
}


def _resolve_model(model_ref):
    """Resolve a model from either a class or 'app.Model' string."""
    if isinstance(model_ref, str):
        from django.apps import apps
        app_label, model_name = model_ref.rsplit('.', 1)
        return apps.get_model(app_label, model_name)
    return model_ref


def _get_backup_files():
    """Return list of backup file metadata dicts, sorted by date descending."""
    if not os.path.exists(BACKUP_DIR):
        return []
    backups = []
    for fname in os.listdir(BACKUP_DIR):
        if not fname.endswith('.json'):
            continue
        fpath = os.path.join(BACKUP_DIR, fname)
        try:
            with open(fpath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            meta = data.get('metadata', {})
            size_bytes = os.path.getsize(fpath)
            if size_bytes < 1024:
                size_str = f"{size_bytes} B"
            elif size_bytes < 1048576:
                size_str = f"{size_bytes / 1024:.1f} KB"
            else:
                size_str = f"{size_bytes / 1048576:.1f} MB"
            backups.append({
                'filename': fname,
                'created_at': meta.get('created_at', ''),
                'type': meta.get('type', 'unknown'),
                'description': meta.get('description', ''),
                'modules': meta.get('modules', []),
                'record_count': meta.get('record_count', 0),
                'schema_version': meta.get('schema_version', ''),
                'size': size_str,
                'size_bytes': size_bytes,
            })
        except (json.JSONDecodeError, OSError):
            backups.append({
                'filename': fname,
                'created_at': '',
                'type': 'corrupto',
                'description': 'Archivo no legible',
                'modules': [],
                'record_count': 0,
                'schema_version': '',
                'size': '?',
                'size_bytes': 0,
            })
    backups.sort(key=lambda b: b['created_at'], reverse=True)
    return backups


@has_permission('view_backups')
def admin_backups(request):
    from ecosite.models import BackupSettings
    backups = _get_backup_files()
    backup_settings = BackupSettings.load()
    pending_orders_count = Order.objects.real().filter(payment_status__in=["review", "paid", "failed"], status__in=['New', 'Processing']).count()
    context = {
        'active_page': 'backups',
        'backups': backups,
        'modules': list(BACKUP_MODULES.keys()),
        'pending_orders_count': pending_orders_count,
        'backup_settings': backup_settings,
    }
    return render(request, 'admin_panel/backups.html', context)


@has_permission('manage_backups')
def admin_create_backup(request):
    if request.method != 'POST':
        return redirect('admin_backups')

    backup_type = request.POST.get('backup_type', 'full')
    description = request.POST.get('description', '').strip()
    selected_modules = request.POST.getlist('modules')

    if backup_type == 'full':
        selected_modules = list(BACKUP_MODULES.keys())
    elif not selected_modules:
        messages.error(request, 'Debes seleccionar al menos un módulo para un respaldo selectivo.')
        return redirect('admin_backups')

    # Collect data
    backup_data = {}
    total_records = 0
    tables_list = []

    for mod_name in selected_modules:
        if mod_name not in BACKUP_MODULES:
            continue
        for model_ref in BACKUP_MODULES[mod_name]:
            model_cls = _resolve_model(model_ref)
            db_table = model_cls._meta.db_table
            tables_list.append(db_table)
            queryset = model_cls.objects.all()
            serialized = serializers.serialize('json', queryset)
            backup_data[db_table] = json.loads(serialized)
            total_records += queryset.count()

    now = timezone.now()
    timestamp_str = now.strftime('%Y%m%d_%H%M%S')
    type_label = 'completo' if backup_type == 'full' else 'parcial'
    filename = f"backup_{type_label}_{timestamp_str}.json"

    backup_json = {
        'metadata': {
            'schema_version': SCHEMA_VERSION,
            'created_at': now.isoformat(),
            'type': type_label,
            'description': description or f'Respaldo {type_label}',
            'modules': selected_modules,
            'tables': tables_list,
            'record_count': total_records,
            'django_version': django.get_version(),
            'app': 'IAAStore',
        },
        'data': backup_data,
    }

    os.makedirs(BACKUP_DIR, exist_ok=True)
    fpath = os.path.join(BACKUP_DIR, filename)
    with open(fpath, 'w', encoding='utf-8') as f:
        json.dump(backup_json, f, ensure_ascii=False, indent=2)

    log_activity(request, 'CREATE', 'Respaldos',
                 f'Creo un respaldo {type_label}: "{filename}" ({total_records} registros). Modulos: {", ".join(selected_modules)}.')
    messages.success(request, f'Respaldo "{filename}" creado exitosamente con {total_records} registros.')
    return redirect('admin_backups')


@has_permission('view_backups')
def admin_download_backup(request, filename):
    fpath = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(fpath) or '..' in filename:
        messages.error(request, 'Archivo de respaldo no encontrado.')
        return redirect('admin_backups')
    return FileResponse(open(fpath, 'rb'), as_attachment=True, filename=filename)


@has_permission('manage_backups')
def admin_restore_backup(request, filename):
    if request.method != 'POST':
        return redirect('admin_backups')

    fpath = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(fpath) or '..' in filename:
        messages.error(request, 'Archivo de respaldo no encontrado.')
        return redirect('admin_backups')

    try:
        with open(fpath, 'r', encoding='utf-8') as f:
            backup_json = json.load(f)
    except json.JSONDecodeError:
        messages.error(request, 'El archivo de respaldo esta corrupto o no es un JSON valido.')
        return redirect('admin_backups')

    # Validate structure
    if 'metadata' not in backup_json or 'data' not in backup_json:
        messages.error(request, 'Estructura del respaldo invalida: faltan campos "metadata" o "data".')
        return redirect('admin_backups')

    metadata = backup_json['metadata']
    data = backup_json['data']

    # Validate schema version
    file_schema = metadata.get('schema_version', '')
    if file_schema and file_schema != SCHEMA_VERSION:
        messages.error(request, f'Version de esquema incompatible: archivo={file_schema}, sistema={SCHEMA_VERSION}.')
        return redirect('admin_backups')

    restored_count = 0
    try:
        with transaction.atomic():
            for table_name, records_json in data.items():
                if not records_json:
                    continue
                # Convert back to JSON string for deserialize
                records_str = json.dumps(records_json)
                objects = list(serializers.deserialize('json', records_str))
                if not objects:
                    continue
                # Delete existing records of this model
                model_cls = objects[0].object.__class__
                model_cls.objects.all().delete()
                # Save all objects
                for obj in objects:
                    obj.save()
                    restored_count += 1

        modules_str = ', '.join(metadata.get('modules', []))
        log_activity(request, 'UPDATE', 'Respaldos',
                     f'Restauro la base de datos desde "{filename}" ({restored_count} registros). Modulos: {modules_str}.')
        messages.success(request, f'Base de datos restaurada exitosamente desde "{filename}" ({restored_count} registros).')
    except Exception as e:
        messages.error(request, f'Error al restaurar: {str(e)}. Los cambios fueron revertidos automaticamente.')

    return redirect('admin_backups')


@has_permission('manage_backups')
def admin_delete_backup(request, filename):
    if request.method != 'POST':
        return redirect('admin_backups')

    fpath = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(fpath) or '..' in filename:
        messages.error(request, 'Archivo de respaldo no encontrado.')
        return redirect('admin_backups')

    try:
        os.remove(fpath)
        log_activity(request, 'DELETE', 'Respaldos', f'Elimino el respaldo "{filename}".')
        messages.success(request, f'Respaldo "{filename}" eliminado.')
    except OSError as e:
        messages.error(request, f'Error al eliminar el archivo: {str(e)}')

    return redirect('admin_backups')


@has_permission('manage_backups')
def admin_upload_backup(request):
    if request.method != 'POST':
        return redirect('admin_backups')

    uploaded = request.FILES.get('backup_file')
    if not uploaded:
        messages.error(request, 'No se selecciono ningun archivo.')
        return redirect('admin_backups')

    if not uploaded.name.endswith('.json'):
        messages.error(request, 'Solo se permiten archivos con extension .json.')
        return redirect('admin_backups')

    # Read and validate
    try:
        content = uploaded.read().decode('utf-8')
        backup_json = json.loads(content)
    except (json.JSONDecodeError, UnicodeDecodeError):
        messages.error(request, 'El archivo subido no es un JSON valido.')
        return redirect('admin_backups')

    if 'metadata' not in backup_json or 'data' not in backup_json:
        messages.error(request, 'Estructura del archivo invalida: faltan campos "metadata" o "data".')
        return redirect('admin_backups')

    # Save file
    os.makedirs(BACKUP_DIR, exist_ok=True)
    base_name = uploaded.name.replace('.json', '')
    timestamp_str = timezone.now().strftime('%Y%m%d_%H%M%S')
    safe_name = f"{base_name}_{timestamp_str}.json"
    fpath = os.path.join(BACKUP_DIR, safe_name)

    with open(fpath, 'w', encoding='utf-8') as f:
        f.write(content)

    record_count = backup_json.get('metadata', {}).get('record_count', '?')
    log_activity(request, 'CREATE', 'Respaldos',
                 f'Subio un archivo de respaldo externo: "{safe_name}" ({record_count} registros).')
    messages.success(request, f'Archivo "{safe_name}" subido exitosamente. Ahora puedes restaurarlo desde la lista.')
    return redirect('admin_backups')


@has_permission('manage_backups')
def admin_update_backup_schedule(request):
    """Actualiza la configuracion de respaldo automatico desde el panel."""
    if request.method != 'POST':
        return redirect('admin_backups')

    from ecosite.models import BackupSettings

    config = BackupSettings.load()
    enabled = request.POST.get('backup_enabled') == 'on'
    interval = request.POST.get('interval_minutes', '1440')

    try:
        interval = int(interval)
        if interval < 1:
            interval = 1
    except (ValueError, TypeError):
        interval = 1440

    config.enabled = enabled
    config.interval_minutes = interval
    config.save()  # Esto dispara el signal post_save que actualiza el scheduler

    estado = 'activado' if enabled else 'desactivado'
    log_activity(request, 'UPDATE', 'Respaldos',
                 f'Actualizo la programacion de respaldos: {estado}, cada {interval} minutos.')
    messages.success(request, f'Programacion de respaldo automatico {estado} (cada {interval} min).')
    return redirect('admin_backups')

# ==========================================
# EXCEL IMPORT / EXPORT
# ==========================================
@has_permission('view_inventory')
def admin_export_products(request):
    import openpyxl
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos_exportados.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    columns = [
        "Código", "Producto", "Categoría", "Cantidad", "Precio $", "Peso kg",
        "Largo", "cm/m", "Ancho", "cm/m", "Alto", "cm/m"
    ]
    ws.append(columns)
    
    # Set reasonable column widths
    column_widths = {
        'A': 20, # Código
        'B': 40, # Producto
        'C': 15, # Categoría
        'D': 10, # Cantidad
        'E': 10, # Precio
        'F': 10, # Peso
        'G': 10, # Largo
        'H': 8,  # cm/m
        'I': 10, # Ancho
        'J': 8,  # cm/m
        'K': 10, # Alto
        'L': 8   # cm/m
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    from django.db.models import Q
    
    products = Product.objects.select_related('Category').all()
    
    q = request.GET.get('q', '').lower()
    c = request.GET.get('c', '')
    s = request.GET.get('s', '')
    min_price = request.GET.get('min', '')
    max_price = request.GET.get('max', '')
    
    if q:
        products = products.filter(
            Q(product_name__icontains=q) | 
            Q(sku__icontains=q)
        )
    if c:
        products = products.filter(Category__category_name__iexact=c)
    if s == 'available':
        products = products.filter(is_available=True, stock__gt=0)
    elif s == 'out_of_stock':
        products = products.filter(stock__lte=0)
    elif s == 'inactive':
        products = products.filter(is_available=False)
        
    if min_price:
        try:
            products = products.filter(price__gte=float(min_price))
        except ValueError:
            pass
    if max_price:
        try:
            products = products.filter(price__lte=float(max_price))
        except ValueError:
            pass
            
    for p in products:
        ws.append([
            p.sku,
            p.product_name,
            p.Category.category_name if p.Category else "",
            p.stock,
            p.price,
            p.gross_weight,
            p.length,
            p.length_unit,
            p.width,
            p.width_unit,
            p.height,
            p.height_unit
        ])
        
    wb.save(response)
    return response

@has_permission('edit_inventory')
def admin_import_products(request):
    import openpyxl
    from django.utils.text import slugify
    from category.models import Category
    from store.models import Product
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            error_count = 0
            
            headers = [cell.value for cell in ws[1]]
            def get_val(row, col_name):
                try:
                    idx = headers.index(col_name)
                    return row[idx].value
                except ValueError:
                    return None
            
            for row in ws.iter_rows(min_row=2):
                if all(cell.value is None for cell in row):
                    continue
                    
                codigo = get_val(row, "Código")
                producto = get_val(row, "Producto")
                categoria_name = get_val(row, "Categoría")
                cantidad = get_val(row, "Cantidad")
                precio = get_val(row, "Precio $")
                peso = get_val(row, "Peso kg")
                largo = get_val(row, "Largo")
                largo_unit = get_val(row, "cm/m")
                ancho = get_val(row, "Ancho")
                ancho_unit = get_val(row, "cm/m")
                alto = get_val(row, "Alto")
                # Handle second or third cm/m (we can rely on specific indices, but this generic get_val returns the first match)
                
                # To accurately get the units since there are three "cm/m" columns, let's index manually
                try:
                    idx_codigo = headers.index("Código")
                    idx_producto = headers.index("Producto")
                    idx_categoria = headers.index("Categoría")
                    idx_cantidad = headers.index("Cantidad")
                    idx_precio = headers.index("Precio $")
                    idx_peso = headers.index("Peso kg")
                    idx_largo = headers.index("Largo")
                    idx_largo_unit = idx_largo + 1
                    idx_ancho = headers.index("Ancho")
                    idx_ancho_unit = idx_ancho + 1
                    idx_alto = headers.index("Alto")
                    idx_alto_unit = idx_alto + 1
                    
                    codigo = row[idx_codigo].value
                    producto = row[idx_producto].value
                    categoria_name = row[idx_categoria].value
                    cantidad = row[idx_cantidad].value
                    precio = row[idx_precio].value
                    peso = row[idx_peso].value
                    largo = row[idx_largo].value
                    largo_unit = row[idx_largo_unit].value
                    ancho = row[idx_ancho].value
                    ancho_unit = row[idx_ancho_unit].value
                    alto = row[idx_alto].value
                    alto_unit = row[idx_alto_unit].value
                except Exception:
                    error_count += 1
                    continue
                
                if codigo:
                    try:
                        p = Product.objects.get(sku=codigo)
                        if cantidad is not None:
                            try: p.stock = int(cantidad)
                            except ValueError: pass
                        if precio is not None:
                            try: p.price = int(precio)
                            except ValueError: pass
                        if peso is not None:
                            try: p.gross_weight = float(peso)
                            except ValueError: pass
                        if largo is not None:
                            try: p.length = float(largo)
                            except ValueError: pass
                        if largo_unit in ['cm', 'm']:
                            p.length_unit = largo_unit
                        if ancho is not None:
                            try: p.width = float(ancho)
                            except ValueError: pass
                        if ancho_unit in ['cm', 'm']:
                            p.width_unit = ancho_unit
                        if alto is not None:
                            try: p.height = float(alto)
                            except ValueError: pass
                        if alto_unit in ['cm', 'm']:
                            p.height_unit = alto_unit
                        
                        p.save()
                        updated_count += 1
                    except Product.DoesNotExist:
                        error_count += 1
                else:
                    if not producto or not categoria_name or precio is None or cantidad is None:
                        error_count += 1
                        continue
                        
                    try:
                        cat = Category.objects.get(category_name=categoria_name)
                    except Category.DoesNotExist:
                        error_count += 1
                        continue
                        
                    slug = slugify(producto)
                    base_slug = slug
                    counter = 1
                    while Product.objects.filter(slug=slug).exists():
                        slug = f"{base_slug}-{counter}"
                        counter += 1
                        
                    p = Product(
                        product_name=producto,
                        slug=slug,
                        Category=cat,
                        price=int(precio),
                        stock=int(cantidad),
                        is_available=False
                    )
                    
                    if peso is not None:
                        try: p.gross_weight = float(peso)
                        except ValueError: pass
                    if largo is not None:
                        try: p.length = float(largo)
                        except ValueError: pass
                    if largo_unit in ['cm', 'm']: p.length_unit = largo_unit
                    if ancho is not None:
                        try: p.width = float(ancho)
                        except ValueError: pass
                    if ancho_unit in ['cm', 'm']: p.width_unit = ancho_unit
                    if alto is not None:
                        try: p.height = float(alto)
                        except ValueError: pass
                    if alto_unit in ['cm', 'm']: p.height_unit = alto_unit
                    
                    p.save()
                    created_count += 1
                    
            messages.success(request, f'Importación completada. Actualizados: {updated_count}, Creados: {created_count}. Errores: {error_count}')
            log_activity(request, 'IMPORT', 'Productos', f'Importación por Excel. Creados: {created_count}, Actualizados: {updated_count}.')
        except Exception as e:
            messages.error(request, f'Error al procesar el Excel: {str(e)}')
            
    return redirect('admin_products')

import json
from django.http import JsonResponse
from .admin_chatbot_engine import AdminChatEngine

def admin_chat_message(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    message = data.get('message', '').strip()
    path = data.get('path', '').strip()

    if not message:
        return JsonResponse({'error': 'Empty message'}, status=400)

    engine = AdminChatEngine()
    response = engine.process(message, request.user, path)

    return JsonResponse(response)
